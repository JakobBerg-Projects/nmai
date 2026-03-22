"""Extract text and image data from base64-encoded file attachments."""

from __future__ import annotations

import base64
import csv
import io
import logging
from dataclasses import dataclass

logger = logging.getLogger(__name__)


@dataclass
class ExtractedFile:
    filename: str
    text: str | None = None
    image_base64: str | None = None
    mime_type: str | None = None


def process_files(files: list[dict]) -> list[ExtractedFile]:
    """Process a list of file attachments from the competition payload."""
    results: list[ExtractedFile] = []
    for f in files:
        filename = f.get("filename", "unknown")
        content_b64 = f.get("content_base64", "")
        mime = f.get("mime_type", "")

        try:
            raw = base64.b64decode(content_b64)
        except Exception:
            logger.warning("Failed to decode base64 for %s", filename)
            continue

        fname_lower = filename.lower()

        if mime == "application/pdf" or fname_lower.endswith(".pdf"):
            results.append(_extract_pdf(filename, raw))
        elif mime.startswith("image/") or fname_lower.endswith((".png", ".jpg", ".jpeg", ".gif", ".webp")):
            results.append(
                ExtractedFile(
                    filename=filename,
                    image_base64=content_b64,
                    mime_type=mime or "image/png",
                )
            )
        elif (
            mime in ("text/csv", "application/csv", "application/vnd.ms-excel")
            or fname_lower.endswith(".csv")
        ):
            results.append(_extract_csv(filename, raw))
        elif fname_lower.endswith((".xlsx", ".xls")) or mime in (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
        ):
            results.append(_extract_excel(filename, raw))
        else:
            text = raw.decode("utf-8", errors="replace")
            results.append(ExtractedFile(filename=filename, text=text))

    return results


def _extract_csv(filename: str, raw_bytes: bytes) -> ExtractedFile:
    """Parse CSV file into a readable table format."""
    try:
        text = raw_bytes.decode("utf-8-sig", errors="replace")  # handles BOM
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        if not rows:
            return ExtractedFile(filename=filename, text="[Empty CSV file]")

        # Format as aligned table
        col_widths = [max(len(str(r[i])) for r in rows if i < len(r)) for i in range(max(len(r) for r in rows))]
        lines = []
        for i, row in enumerate(rows):
            line = " | ".join(str(cell).ljust(col_widths[j]) for j, cell in enumerate(row))
            lines.append(line)
            if i == 0:
                lines.append("-" * len(line))  # separator after header

        formatted = f"[CSV FILE: {filename}]\n" + "\n".join(lines)
        logger.info("Extracted CSV %s: %d rows, %d cols", filename, len(rows), len(rows[0]) if rows else 0)
        return ExtractedFile(filename=filename, text=formatted)
    except Exception as exc:
        logger.warning("CSV extraction failed for %s: %s", filename, exc)
        # Fallback: return raw text
        return ExtractedFile(filename=filename, text=raw_bytes.decode("utf-8", errors="replace"))


def _extract_excel(filename: str, raw_bytes: bytes) -> ExtractedFile:
    """Extract data from Excel files using openpyxl."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
        parts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            parts.append(f"[SHEET: {sheet_name}]")
            # Filter out completely empty rows
            non_empty = [r for r in rows if any(c is not None for c in r)]
            if not non_empty:
                continue
            col_count = max(len(r) for r in non_empty)
            col_widths = [
                max((len(str(r[i])) if i < len(r) and r[i] is not None else 0) for r in non_empty)
                for i in range(col_count)
            ]
            for j, row in enumerate(non_empty):
                cells = [(str(row[i]) if i < len(row) and row[i] is not None else "") for i in range(col_count)]
                line = " | ".join(c.ljust(col_widths[i]) for i, c in enumerate(cells))
                parts.append(line)
                if j == 0:
                    parts.append("-" * len(line))

        text = f"[EXCEL FILE: {filename}]\n" + "\n".join(parts)
        logger.info("Extracted Excel %s: %d sheets", filename, len(wb.sheetnames))
        return ExtractedFile(filename=filename, text=text)
    except ImportError:
        logger.warning("openpyxl not installed, treating Excel as binary")
        return ExtractedFile(filename=filename, text=f"[Excel file {filename} — could not parse: openpyxl not installed]")
    except Exception as exc:
        logger.warning("Excel extraction failed for %s: %s", filename, exc)
        return ExtractedFile(filename=filename, text=f"[Excel file {filename} — extraction failed: {exc}]")


def _extract_pdf(filename: str, raw_bytes: bytes) -> ExtractedFile:
    """Extract text from a PDF. Falls back to first-page image if text is sparse."""
    try:
        import fitz

        doc = fitz.open(stream=raw_bytes, filetype="pdf")
        text_parts: list[str] = []
        for page in doc:
            text_parts.append(page.get_text())
        text = "\n".join(text_parts).strip()

        if len(text) > 50:
            return ExtractedFile(filename=filename, text=text)

        # Sparse text — render first page as image for the LLM's vision
        page = doc[0]
        pix = page.get_pixmap(dpi=200)
        img_bytes = pix.tobytes("png")
        img_b64 = base64.b64encode(img_bytes).decode()
        return ExtractedFile(
            filename=filename,
            text=text if text else None,
            image_base64=img_b64,
            mime_type="image/png",
        )
    except Exception as exc:
        logger.warning("PDF extraction failed for %s: %s", filename, exc)
        return ExtractedFile(filename=filename, text=f"[Could not extract PDF: {exc}]")


def build_file_context(extracted: list[ExtractedFile]) -> str:
    """Build a text summary of all extracted files for the LLM prompt."""
    if not extracted:
        return ""

    parts: list[str] = ["## Attached Files\n"]
    for ef in extracted:
        parts.append(f"### {ef.filename}")
        if ef.text:
            parts.append(ef.text)
        if ef.image_base64 and not ef.text:
            parts.append("[Image attached — see image content in this message]")
        parts.append("")

    return "\n".join(parts)
